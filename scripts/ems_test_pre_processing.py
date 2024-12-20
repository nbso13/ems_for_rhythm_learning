
import warnings
import pickle
import ems_constants
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from ems_test import process_contact_trace_to_hit_times
from ems_test import plot_contact_trace_and_rhythm
import glob
import quantities as pq
from elephant.spike_train_dissimilarity import victor_purpura_distance
from neo import SpikeTrain
import scipy
import numpy as np
import pandas as pd
from scipy import signal
import os
import time
from ems_test_analysis import pull_repeat_times
from ems_test_analysis import plot_each_block
from ems_test_analysis import determine_delays_list
from ems_test_analysis import plot_test_blocks

def sine_generator(fs, sinefreq, duration):
    T = duration
    nsamples = fs * T
    w = 2. * np.pi * sinefreq
    t_sine = np.linspace(0, T, nsamples, endpoint=False)
    y_sine = np.sin(w * t_sine)
    result = pd.DataFrame({ 
        'data' : y_sine} ,index=t_sine)
    return result

def butter_highpass(cutoff, fs, order=5):
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    b, a = signal.butter(order, normal_cutoff, btype='high', analog=False)
    return b, a

def butter_lowpass(cutOff, fs, order=5):
    nyq = 0.5 * fs
    normalCutoff = cutOff / nyq
    b, a = signal.butter(order, normalCutoff, btype='low', analog = True)
    return b, a

def butter_lowpass_filter(data, cutOff, fs, order=4):
    b, a = butter_lowpass(cutOff, fs, order=order)
    y = signal.lfilter(b, a, data)
    return y

def butter_band_pass_filter(x_vec, data, cutoff_low, cutoff_high, fs, plot_flag, order=5):
    b, a = butter_highpass(cutoff_low, fs, order=order)
    y = signal.filtfilt(b, a, data)
    # y = butter_lowpass_filter(y, cutoff_high, fs)
    if plot_flag == 1:
        plt.figure(figsize=(20,10))
        plt.subplot(211)
        plt.plot(x_vec, data)
        plt.title('input signal')
        plt.subplot(212)
        plt.plot(x_vec, y)
        plt.title('filtered signal')
        plt.show()
    return y

def spike_times_to_traces(onset_times, hold_length, x_vector, samp_period):
    # take a series of onset time points and craft plottable traces.
    ### XVECTOR SHOULD HAVE PERIOD OF SAMP PERIOD
    array_value_stim_time = int(np.floor(hold_length/samp_period))
    if array_value_stim_time < 2:
        array_value_stim_time = 2
    trace = np.zeros_like(x_vector)
    for time_val in onset_times: # for each onset time
        array_ind_begin = int(np.floor((time_val-x_vector[0])/samp_period)) 
        array_ind_end = array_ind_begin + array_value_stim_time
        trace[array_ind_begin:array_ind_end] = 1
    return trace

def plot_w_k(axes_array, means, sds, title_list):
    axes = sum([list(axes_array[0]), list(axes_array[1])], [])
    slopes = []
    intercepts = []
    r_vals = []
    p_vals = []
    for i in range(len(means)):
        slope, intercept, r_value, p_value, std_err = scipy.stats.linregress(means[i], sds[i])
        slopes.append(slope)
        intercepts.append(intercept)
        r_vals.append(r_value)
        p_vals.append(p_value)
        axes[i].scatter(means[i], sds[i])
        axes[i].plot(means[i], slope * np.array(means[i]) + intercept)
        axes[i].set_title(title_list[i])
        axes[i].text(min(means[i]), max(sds[i]), "R^2: " + str(r_value**2))
        axes[i].set_ylabel("Standard deviation (ms)")
        axes[i].set_xlabel("Mean Produced Interval (ms)")
    plt.ion()
    plt.tight_layout()
    plt.show()
    plt.draw()
    plt.pause(0.01)
    return slopes, intercepts, r_vals, p_vals

def interpolate(reading_list, contact_x, samp_period):
    contact_x = contact_x - contact_x[0]
    contact_x_interped = np.arange(0, max(contact_x), samp_period)
    f = scipy.interpolate.interp1d(contact_x, reading_list)
    reading_interped = f(contact_x_interped)
    return contact_x_interped, reading_interped



def plot_traces(x_array, trace_list, samp_period, legend_labels, title):
    fig, ax = plt.subplots()
    ax.set_yticks(np.arange(0, 500, 100))
    ax.set_xticks(np.arange(np.min(x_array), np.max(x_array), 1000))
    ax.plot(x_array, trace_list[0])
    for i in range(len(trace_list) - 1):
        ax.plot(x_array, trace_list[i+1]*np.max(trace_list[0])/2)
    ax.legend(legend_labels)
    ax.set_title(title)
    ax.set_xlabel("time, ms")
    ax.set_ylabel("onset")
    plt.ion()
    plt.show()
    plt.draw()
    plt.pause(0.01)

def earth_movers_distance(spike_times_a, spike_times_b, rhythm_trace_a, rhythm_trace_b):
    
    rhythm_total_spike_times_a = len(spike_times_a)
    rhythm_total_spike_times_b = len(spike_times_b)
    if rhythm_total_spike_times_a == 0:
        return -1
    cumulative_a = np.cumsum(np.divide(rhythm_trace_a, rhythm_total_spike_times_a))
    cumulative_b = np.cumsum(np.divide(rhythm_trace_b, rhythm_total_spike_times_b))
    # same thing as np.sum(np.abs(np.subtract(cumulative_a, cumulative_b))),
    return  scipy.stats.wasserstein_distance(spike_times_a, spike_times_b)


# TEST THIS FUNCTION
def compile_unique_interval_list(intervals, tolerance):
    interval_index_lib = np.arange(0, len(intervals))
    unique_interval_list = []
    index = 0
    indices = []
    for interval in intervals: # for every interval
        # if this interval is within tolerance of any item on the list,
        bool_list = [(interval+tolerance > item and interval-tolerance < item) for item in unique_interval_list]
        if any(bool_list):
            indices.append(bool_list.index(True))
            continue # we consider it on the list.
        else: #otherwise
            unique_interval_list.append(interval) # add it to list
            indices.append(len(unique_interval_list) - 1)
    return unique_interval_list, indices

def count_time_readings(worksheet):
## how many time readings are there?
    counter = 0
    val = 1.0
    while type(val) is float or type(val) is int: #while we're reading floats and not nans
        val = worksheet.cell(row=counter + worksheet_data_begin_indices[0], column=1).value
        counter = counter+1
    time_readings_length = counter-2
    return time_readings_length

def read_array_values(time_readings_length, variables_number, worksheet):
    arr = np.empty([time_readings_length, variables_number]) 
    arr[:] = np.NaN
    for r in range(time_readings_length):
        for c in range(variables_number):
            arr[r][c] = worksheet.cell(row=r + worksheet_data_begin_indices[0], column=c + worksheet_data_begin_indices[1]).value
    contact_x_values = arr[:, 0]
    reading_list = arr[:, 1]
    stim_onsets_temp = arr[:, 2]
    stim_onset_times = stim_onsets_temp[~np.isnan(stim_onsets_temp)] # take care of nans that make array work
    audio_onsets_temp = arr[:, 3]
    audio_onset_times = audio_onsets_temp[~np.isnan(audio_onsets_temp)]
    return contact_x_values, reading_list, stim_onset_times, audio_onset_times

        
def chop_traces(k, surpressed_contact_onset_times, surpressed_contact_trace, audio_onset_times, audio_trace, delays_list, x_values):
    loop_begin = delays_list[k] - 0.5*30000/bpm #include half an eighthnote before
    loop_end = delays_list[k+1] + 1 * 30000/bpm #include half an eighthnote after as well
    contact_bool = np.logical_and((surpressed_contact_onset_times >= loop_begin), (surpressed_contact_onset_times <= loop_end)) # select contact onset times during this loop of rhythm
    audio_bool = np.logical_and((audio_onset_times >= loop_begin), (audio_onset_times <= loop_end)) # select audio onset times during this loop of rhythm
    spike_times_contact = surpressed_contact_onset_times[contact_bool] - loop_begin # how many spikes total?
    spike_times_audio = audio_onset_times[audio_bool] - loop_begin
    trace_selector_bool = np.logical_and((x_values >= loop_begin), (x_values <= loop_end)) # which indices in traces are during this loop?
    contact_trace_selected = surpressed_contact_trace[trace_selector_bool] # pick those data points from suprpressed contact trace
    audio_trace_selected = audio_trace[trace_selector_bool] # pick those data points from audio trace
    return spike_times_contact, spike_times_audio, contact_trace_selected, audio_trace_selected, trace_selector_bool

def load_headers(file_stems):
    header_dicts = []
    for file_stem in file_stems:
        pkl_file = f"data/{file_stem}_header_info.pkl"
        with open(pkl_file, "rb") as pkl_handle:
            header_dict = pickle.load(pkl_handle)
        header_dicts.append(header_dict)
    return header_dicts

def histo_intervals(gt_intervals):
    # q25, q75 = np.percentile(gt_intervals, [25, 75])
    # bin_width = 2 * (q75 - q25) * len(gt_intervals) ** (-1/3)
    # bins = round((np.max(gt_intervals) - np.min(gt_intervals)) / bin_width)
    fig, ax = plt.subplots()
    ax.hist(gt_intervals, bins=40)  # density=False would make counts
    ax.set_ylabel("count")
    ax.set_xlabel('Interval length, ms')
    ax.set_title("interval frequency")
    return

def load_data(file_stems):
    wbs = []
    for file_stem in file_stems:
        print(f"loading {file_stem}...")
        xls_file = f"data/{file_stem}.xlsx"
        wb = load_workbook(xls_file) # most recent participant file
        print("done.")
        wbs.append(wb)
    return wbs

def plot_emds(emds, header_dict, rhyth_index):
    mean_distances = np.mean(np.vstack(emds), 0)
    sd_distances = np.std(np.vstack(emds), 0)
    fig, ax = plt.subplots() # change this according to num phase?
    legend_labels = ["earth mover's distances mean", "std+", "std-"]
    ax.plot(np.arange(len(mean_distances)), mean_distances,'b')
    ax.plot(np.arange(len(mean_distances)), mean_distances+sd_distances, 'r')
    ax.plot(np.arange(len(mean_distances)), mean_distances-sd_distances, 'r')
    ax.set_title(f"mean earth mover's distance across bpms for each phase, {header_dict['rhythm_strings_names'][rhyth_index]}")
    ax.legend(legend_labels)
    plt.ion()
    plt.show()
    plt.draw()
    plt.pause(0.01)
    return

def craft_x_vec(list_of_onsets):
    mini = np.min(list_of_onsets)
    maxi = np.max(list_of_onsets)
    stepsize = (maxi-mini)/(len(list_of_onsets))
    return np.arange(mini, maxi, stepsize)

def count_intervals(rhyth_string):
    zero_counter = 0
    intervs = []
    for i in range(len(rhyth_string)):
        if rhyth_string[i] == '0':
            zero_counter += 1 # add to interval
        if rhyth_string[i] == '1':
            intervs.append(zero_counter) # save interval
            zero_counter = 0 # reset counter
    if zero_counter == len(rhyth_string): # no intervals, empty string
        unique_intervals = []
        num_unique = 0
        return intervs, unique_intervals, num_unique
    intervs[0] += zero_counter # add trailing zeros to first measured interval
    unique_intervals = list(set(intervs))
    num_unique = len(unique_intervals)
    return intervs, unique_intervals, num_unique
    

def emd_test(title, times_a, times_b, mini, maxi, samp_period):
    xvec = np.arange(mini, maxi, samp_period)
    trace_a = spike_times_to_traces(times_a, samp_period, xvec, samp_period)
    trace_b = spike_times_to_traces(times_b, samp_period, xvec, samp_period)
    emd = earth_movers_distance(times_a, times_b, trace_a, trace_b)
    title = title + f", emd = {emd}"
    plot_traces(xvec, [trace_a, trace_b], samp_period, ["a", "b"], title)
    return emd

def filter_test():
    fps = 30
    sine_fq = 10 #Hz
    duration = 10 #seconds
    sine_5Hz = sine_generator(fps,sine_fq,duration)
    sine_fq = 1 #Hz
    duration = 10 #seconds
    sine_1Hz = sine_generator(fps,sine_fq,duration)
    sine_fq = 50 #Hz
    duration = 10 #seconds
    sine_50Hz = sine_generator(fps,sine_fq,duration)

    sine = sine_5Hz + sine_1Hz + sine_50Hz
    x = np.arange(0, duration, 1/fps)
    filtered_sine = butter_band_pass_filter(x, sine.data, 9, 20, fps, plot_flag=1)



def teaserfigure():
    mini = 0
    maxi = 6
    samp_period = 0.01
    # titles = ["times b just after", "times b just before", "times b mixed", \
    #     "times b missing beat", "times b missing two beats", "times b extra beat", \
    #         "times b extra two beats", "times b extra beat end range", "times b extra beat mid range"]

    # times_a = [1, 2, 3, 4, 5]
    # times_bs = [ [1.1, 2.1, 3.1, 4.1, 5.1], [0.9, 1.9, 2.9, 3.9, 4.9], [1.1, 1.9, 3.1, 3.9, 5.1], \
    #     [1.1, 2.1, 3.1, 5.1], [1.1, 3.1, 5.1], [1.1, 2.1, 3.1, 4.1, 4.5, 5.1],  \
    #         [1.1, 2.1, 3.1, 3.5, 4.1, 4.5, 5.1 ], [1.1, 2.1, 3.1, 4.1, 4.9, 5.1], [1.1, 2.1, 3.1, 3.2, 4.1, 5.1]]

    titles = ["Before EMS", "During EMS", "After EMS"]

    times_a = [1, 2, 3, 4, 5]
    times_bs = [ [0.5, 0.8, 2.2, 2.9, 3.7, 3.8, 5.1], [1, 2, 3.1, 4, 4.9], [0.8, 2.3, 3.1, 3.8, 5.1]]

    # legend_labels = ["Subject performance", "Target pattern"]
    fig, axes = plt.subplots(1, 4)
    emds = []

    hold_length = 0.05 # ms
    for i in range(len(times_bs)):
        times_b = times_bs[i]
        # title = titles[i]
        
        xvec = np.arange(mini, maxi, samp_period)
        trace_a = spike_times_to_traces(times_a, hold_length, xvec, samp_period)
        trace_b = spike_times_to_traces(times_b, hold_length, xvec, samp_period)
        emds.append(earth_movers_distance(times_a, times_b, trace_a, trace_b))
        axes[i].set_yticks([])
        axes[i].set_xticks([])
        axes[i].plot(xvec, trace_b+1)
        axes[i].plot(xvec, trace_a)
        if i == 0:
            # axes[i].legend(legend_labels, loc='best')
            axes[i].set_yticks([0.5, 1.5])
            axes[i].set_yticklabels(['ground truth \n pattern', 'user produced \n pattern'], fontsize=10)
        axes[i].set_title(titles[i])
        axes[i].set_xlabel("time")


    
    axes[i+1].plot([0,3,6], emds, '-o', color='green')
    axes[i+1].set_yticks([0.05, 0.4])
    axes[i+1].set_yticklabels(['high \nprecision', 'low \nprecision'], fontsize=10)
    axes[i+1].set_xticks([0,3,6])
    axes[i+1].set_xticklabels(titles)
    # plt.tight_layout()
    plt.xticks(rotation=10, ha="right")
    axes[i+1].set_title("Temporal precision by phase")
    l, b, w, h = axes[i+1].get_position().bounds
    axes[i+1].set_position([l*1.05, b, w, h])



def emd_tests():
    mini = 0
    maxi = 6
    samp_period = 0.01
    titles = ["times b just after", "times b just before", "times b mixed", \
        "times b missing beat", "times b missing two beats", "times b extra beat", \
            "times b extra two beats", "times b extra beat end range", "times b extra beat mid range"]

    times_a = [1, 2, 3, 4, 5]
    times_bs = [ [1.1, 2.1, 3.1, 4.1, 5.1], [0.9, 1.9, 2.9, 3.9, 4.9], [1.1, 1.9, 3.1, 3.9, 5.1], \
        [1.1, 2.1, 3.1, 5.1], [1.1, 3.1, 5.1], [1.1, 2.1, 3.1, 4.1, 4.5, 5.1],  \
            [1.1, 2.1, 3.1, 3.5, 4.1, 4.5, 5.1 ], [1.1, 2.1, 3.1, 4.1, 4.9, 5.1], [1.1, 2.1, 3.1, 3.2, 4.1, 5.1]]

    emds = []
    for i in range(len(times_bs)):
        times_b = times_bs[i]
        title = titles[i]
        emds.append(emd_test(title, times_a, times_b, mini, maxi, samp_period))

    fig, ax = plt.subplots()
    ax.set_xticks(np.arange(len(titles)))
    ax.set_xticklabels(titles)
    ax.set_title("Bar plot for EMDS across tests")
    ax.set_ylabel("EMD")
    plt.xticks(rotation=30, ha="right")
    ax.bar(np.arange(len(titles)), emds, align='center')
    return 


def accumulate_intervals(phase_audio_onsets, surpressed_contact_onset_times):
    ground_truth_intervals = []
    user_intervals = []
    user_error = []
    for j in range(len(phase_audio_onsets)-1):
        # get the interval between them
        gt_interval = phase_audio_onsets[j+1] - phase_audio_onsets[j]
        # get the index of the nearest response pulse to the j+1 audio
        arg_min = np.argmin(np.abs(np.subtract(surpressed_contact_onset_times, phase_audio_onsets[j+1]))) # throws out the first one...
        # get the user interval
        user_interval = surpressed_contact_onset_times[arg_min] - phase_audio_onsets[j]#response time user - previous audio pulse
        if user_interval <= 0:
            user_interval = np.nan
            # raise ValueError("user interval less than 0")
        ground_truth_intervals.append(gt_interval)
        user_intervals.append(user_interval)
        user_error.append(np.abs(gt_interval-user_interval))
    return ground_truth_intervals, user_intervals, user_error

def get_all_intervals(header_dict, audio_onset_times, delays_list, surpressed_contact_onset_times):
    var_lists = []
    for k in range(header_dict['num_phases']): # for every phase
        # array of 1s for audio onset times in this phase
        this_phase_bools = np.logical_and((audio_onset_times > delays_list[k]), (audio_onset_times < delays_list[k+1])) 
        #contact trace selection
        this_phase_cont_bools = np.logical_and((surpressed_contact_onset_times > delays_list[k]), (surpressed_contact_onset_times < delays_list[k+1]))
        # get those onset times
        phase_audio_onsets = audio_onset_times[this_phase_bools]
        phase_contact_onsets = surpressed_contact_onset_times[this_phase_cont_bools]
        total_onsets = np.hstack([phase_audio_onsets, phase_contact_onsets])
        # xvec = np.arange(np.min(total_onsets), np.max(total_onsets), header_dict['samp_period_ms'])
        #plot the phase
        # contact_trace = spike_times_to_traces(phase_contact_onsets, 10, xvec, header_dict['samp_period_ms'])
        # audio_trace = spike_times_to_traces(phase_audio_onsets, audio_hold, xvec, header_dict['samp_period_ms'])
        # labels = ["audio", "contact"]
        # title = f"{header_dict['rhythm_strings_names'][rhythm_index]}, {header_dict['bpms'][bpm_index]} bpm, {header_dict['phase_name_strs'][k]}"
        # plot_traces(xvec, [audio_trace, contact_trace], header_dict['samp_period_ms'], labels, title)

        # for each onset time
        ground_truth_intervals, user_intervals, user_error = accumulate_intervals(phase_audio_onsets, surpressed_contact_onset_times)
        var_list = [ground_truth_intervals, user_intervals, user_error]
        var_lists.append(var_list) # now has each WK relevant variable for each phase
    return var_lists

def surpress(audio_trace, surpressed_contact_trace):
# change to pure spikes (complete surround surpression)
    audio_trace_copy = np.copy(audio_trace)
    for j in range(len(audio_trace)-1):
        if audio_trace_copy[j] == 1:
            audio_trace[j+1] = 0
    
    contact_trace_copy = np.copy(surpressed_contact_trace)
    for j in range(len(surpressed_contact_trace)-1):
        if contact_trace_copy[j] == 1:
            surpressed_contact_trace[j+1] = 0
    return surpressed_contact_trace, audio_trace



def bar_plot_scores(names, scores, errors, title, ylabel):
    fig, ax = plt.subplots()
    ax.set_xticks(np.arange(len(scores)))
    ax.set_xticklabels(names)
    ax.set_title(title)
    ax.set_ylabel(ylabel)
    plt.xticks(rotation=45, ha="right")
    ax.bar(np.arange(len(scores)), scores, yerr=errors, align='center')
    plt.tight_layout()

def trace_surpress(reading_list, x_times, memory_ms): # for every time point along the trace,
    # if the mean of the past memory ms is above the threshold, contact was quite recent and this index should be 0'd
    x_times_array = np.array(x_times)
    reading_list_array = np.array(reading_list)
    more_than_thresh_bool = reading_list_array > ems_constants.baseline_subtractor
    indices = np.arange(len(x_times_array))
    indices_selected = indices[more_than_thresh_bool]
    x_times_more_than_thresh = x_times_array[more_than_thresh_bool]
    for time_index in range(len(x_times_more_than_thresh)):
        # this is as far back in time to go
        past_to_examine = x_times_more_than_thresh[time_index] - memory_ms
        # these are the time indices in that range. might be empty!
        times_in_that_range_bool = np.logical_and((x_times_array > past_to_examine), \
            (x_times_array < x_times_more_than_thresh[time_index]))
        # these are the readings in that range
        readings_in_that_range = reading_list_array[times_in_that_range_bool]
        # if no readings in that ranage continue.
        if len(readings_in_that_range) == 0:
            continue
        else:
            # otherwise, calculate the mean and see if it is above baseline subtractor.
            before_mean = np.mean(readings_in_that_range)
            if before_mean > ems_constants.baseline_subtractor/3:
                # if so, surpress this index for reading_list
                ind = indices_selected[time_index]
                reading_list[ind] = 0
    return reading_list




# def cluster_intervals(num_intervals, gt_intervals)
# _____________________________________________________
# _____________________### MAIN ###____________________

if __name__ == '__main__':

    plt.style.use('ggplot')

    # teaserfigure()

    ### run tests ###
    # emd_tests()
    # filter_test()

    # file_stems =  ['2022_03_24_16_43_14_pp4', '2022_03_25_15_46_43_pp4']
    # file_stems =  ['2022_03_27_13_56_12_pp5']
    file_stems = ['2022_04_13_13_37_03_pp11']
    
    # '2022_04_11_15_34_19_pp11', '2022_04_11_15_56_10_pp11']


    ### load data ###
    print("Loading data...")
    t1 = time.time()
    header_dicts = load_headers(file_stems)
    wbs = load_data(file_stems)
    end = time.time() - t1
    print("Data loaded. time taken: " + str(end))

    vector_of_improvement_matrices = []

    for session_number in range(len(wbs)):
        header_dict = header_dicts[session_number]
        wb = wbs[session_number]

        # pull experimental condition string
        condition = ems_constants.counter_balanced_number_dict[header_dict["counter-balanced-number"]]

        # xlsx parser does not do 0-start counting
        worksheet_data_begin_indices = [val + 1 for val in header_dict["worksheet_data_begin_indices" ]]
        variables_number = 4   # time_readings, contact trace, stim_onsets, audio_onsets

        count = 0

        all_emds = []

        repeat_list = header_dict['phase_repeats_list']
        
        processed_vars_by_rhythm = []

        y_axis_maxes = {
            'mads': 0.5,
            'vads': 0.2,
            'emds': 400,
            'twds' : 100
        }

        y_axis_mins = {
            'mads': 0,
            'vads': 0,
            'emds': 0,
            'twds' : 0
        }

        # doing preprocessing FOR EACH RHYTHM separately.

        rhythm_strings = header_dict['rhythm_strings']
        for rhythm_index in range(len(rhythm_strings)): # for each string
            rhythm_substr = rhythm_strings[rhythm_index]
            count = count + 1

            list_of_WK_var_lists_by_bpm = []

            processed_var_lists_by_bpm = []
            processed_vars_by_rhythm.append(processed_var_lists_by_bpm)

            bpms = header_dict['bpms'] # for each bpm
            for bpm_index in range(len(bpms)):
                
                bpm = bpms[bpm_index]
                worksheet = wb[f"{header_dict['rhythm_strings_names'][rhythm_index]}_bpm{bpm}"]
                time_readings_length = count_time_readings(worksheet)

                ## make the appropriate size data array and read in values.
                contact_x_values, reading_list, stim_onset_times, audio_onset_times = read_array_values(time_readings_length, variables_number, worksheet)
                cutoff_freq_low = ems_constants.cutoff_freq_low # cutoff period Any component with a longer period is attenuated
                cutoff_freq_high = ems_constants.cutoff_freq_high
                contact_x_interped, reading_list_interped = interpolate(reading_list, contact_x_values, ems_constants.analysis_sample_period)
                reading_list_interped = trace_surpress(reading_list_interped, contact_x_interped, ems_constants.memory_ms)
                # reading_list_filtered = butter_band_pass_filter(contact_x_interped, reading_list_interped, cutoff_freq_low, cutoff_freq_high, header_dict['samp_period_ms'], plot_flag=0)
            
                ## preprocessing
                audio_hold = 30000/bpm 
                x_vec = contact_x_interped
                stim_trace = spike_times_to_traces(stim_onset_times, header_dict['actual_stim_length'], x_vec, ems_constants.analysis_sample_period)
                audio_trace = spike_times_to_traces(audio_onset_times, audio_hold, x_vec, ems_constants.analysis_sample_period)
                first_audio = audio_onset_times[0]
                last_audio = audio_onset_times[-1]

                # get delays list 
                delays_list, len_rhythm_ms = determine_delays_list(rhythm_substr, bpm, header_dict, first_audio)

                # EXAMINE DATA

                # interpolated
                legend_labels = ['contact', 'stim', 'audio']
                title_str = f"pp {file_stems[session_number][-3:-1]}, session number {session_number}, {header_dict['rhythm_strings_names'][rhythm_index]}, {bpm} interpolated"
                view_window_begin = 40000
                view_window_end = -1
                # plot_contact_trace_and_rhythm(reading_list_interped[view_window_begin:view_window_end], \
                #     x_vec[view_window_begin:view_window_end], stim_trace[view_window_begin:view_window_end],  \
                #         audio_trace[view_window_begin:view_window_end], x_vec[view_window_begin:view_window_end], header_dict['samp_period_ms'], legend_labels, title_str)

                # filtered
                # title_str = f"{header_dict['rhythm_strings_names'][rhythm_index]}, {bpm} filtered"
                # plot_contact_trace_and_rhythm(reading_list_filtered, x_vec, stim_trace, audio_trace, x_vec, header_dict['samp_period_ms'], legend_labels, title_str)

                ## continue preprocessing

                # determine when each contact/hit began
                surpressed_contact_onset_times_not_chopped = process_contact_trace_to_hit_times(reading_list_interped, contact_x_interped, ems_constants.baseline_subtractor, ems_constants.surpression_window)
                # take off onset times that are before or after stim plus or minus 150 ms
                surpressed_contact_onset_times = np.array([time for time in surpressed_contact_onset_times_not_chopped if (time > first_audio-ems_constants.chopping_buffer and time < last_audio+ ems_constants.chopping_buffer)])
                # get the plottable trace for that 
                surpressed_contact_trace = spike_times_to_traces(surpressed_contact_onset_times, ems_constants.contact_spike_time_width, contact_x_interped, ems_constants.analysis_sample_period)

                surpressed_contact_trace, audio_trace = surpress(audio_trace, surpressed_contact_trace)

                repeat_times = pull_repeat_times(first_audio, rhythm_substr, bpm, header_dict['phase_repeats_list'], header_dict['phase_flags_list'])

                # emds, twds, mads, vads = plot_each_block(rhythm_substr, header_dict['rhythm_strings_names'][rhythm_index], bpm, header_dict['phase_repeats_list'], header_dict['phase_flags_list'], delays_list, \
                #     surpressed_contact_onset_times, audio_onset_times, surpressed_contact_trace, audio_trace, x_vec, reading_list, contact_x_values, reading_list_interped, contact_x_interped)

                # y_lims_list = plot_test_blocks(emds, mads, vads, twds, header_dict['rhythm_strings_names'][rhythm_index], rhythm_substr, bpm, ems_constants.phase_warning_strs)

                #examine spiking
                # title_str = f"{header_dict['rhythm_strings_names'][rhythm_index]}, {bpm} spikes"
                # plot_contact_trace_and_rhythm(surpressed_contact_trace[view_window_begin:view_window_end], \
                #     x_vec[view_window_begin:view_window_end], stim_trace[view_window_begin:view_window_end],  \
                #         audio_trace[view_window_begin:view_window_end], x_vec[view_window_begin:view_window_end], header_dict['samp_period_ms'], legend_labels, title_str)


                var_dict = {
                    "contact_x_interped" : contact_x_interped, 
                    "reading_list_interped" : reading_list_interped, 
                    "surpressed_contact_trace" : surpressed_contact_trace, 
                    "surpressed_contact_onset_times" : surpressed_contact_onset_times, 
                    "stim_onset_times" : stim_onset_times, 
                    "audio_onset_times" : audio_onset_times, 
                    "stim_trace" : stim_trace, 
                    "audio_trace" : audio_trace
                }
                processed_var_lists_by_bpm.append(var_dict)
                
                
                ## check delays markers
                # plot_contact_trace_and_rhythm(surpressed_contact_trace, x_vec, stim_trace,  audio_trace, \
                #     x_vec, header_dict['samp_period_ms'], legend_labels, title_str)
                # ax = plt.gca()
                # ax.scatter(delays_list, np.ones_like(delays_list), s=20)

                # check stim accuracy
                if bpm_index == 0 and rhythm_index == 0:
                    x_vec_array = np.array(x_vec)
                    bool_selector = np.logical_and((x_vec_array > delays_list[2]), (x_vec_array < delays_list[3]))
                    title_str = f"{header_dict['rhythm_strings_names'][rhythm_index]}, bpm: {bpm}, stim phase check raw trace"
                    plot_contact_trace_and_rhythm(reading_list_interped[bool_selector], x_vec[bool_selector], stim_trace[bool_selector],  audio_trace[bool_selector], \
                        x_vec[bool_selector], ems_constants.analysis_sample_period, legend_labels, title_str)
                    title_str = f"{header_dict['rhythm_strings_names'][rhythm_index]}, bpm: {bpm}, stim phase check spikes"
                    plot_contact_trace_and_rhythm(surpressed_contact_trace[bool_selector], x_vec[bool_selector], stim_trace[bool_selector],  audio_trace[bool_selector], \
                        x_vec[bool_selector], ems_constants.analysis_sample_period, legend_labels, title_str)


        
            

        print(f"dumping {file_stems[session_number]}.")
        processed_vars_dict = {
            "read_me" : "this dictionary contains the vars_by_rhythm_and_bpm variable which is a list of lists of dictionaries. The first list is by rhythm, the second by bpm. dictionaries contain preprocessed variables for tapping performance and ground truth. header_dict contains metadata such as the rhythms and bpms and participant and time of trial.",
            "vars_by_rhythm_and_bpm" : processed_vars_by_rhythm,
            "header_dict" : header_dict
        }
        with open(f"data/processed_{file_stems[session_number]}.pkl", "wb") as pkl_handle:
            pickle.dump(processed_vars_dict, pkl_handle)
